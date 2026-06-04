import io
from django.http import HttpResponse
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from apps.attendance.models import AttendanceSession, AttendanceRecord
from apps.courses.models import Course, Enrollment

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def course_report_data(request, course_id):
    if not (request.user.is_lecturer or request.user.is_admin):
        return Response({'error':'Access denied'},status=403)
    try:
        course = Course.objects.get(pk=course_id)
    except Course.DoesNotExist:
        return Response({'error':'Course not found'},status=404)
    sessions = AttendanceSession.objects.filter(course=course, is_completed=True).order_by('started_at')
    enrollments = Enrollment.objects.filter(course=course, is_active=True).select_related('student__user')
    rows = []
    for enr in enrollments:
        student = enr.student
        total = sessions.count()
        attended = AttendanceRecord.objects.filter(student=student,course=course,status='PRESENT').count()
        pct = round(attended/total*100,1) if total else 0
        rows.append({'student_id':student.student_id,'name':student.user.get_full_name(),
                     'total_classes':total,'attended':attended,'percentage':pct})
    return Response({'course':course.name,'course_code':course.code,'rows':rows,'total_sessions':sessions.count()})

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_pdf(request, course_id):
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
    except ImportError:
        return HttpResponse('reportlab not installed', status=500)
    try:
        course = Course.objects.get(pk=course_id)
    except Course.DoesNotExist:
        return HttpResponse('Course not found', status=404)
    sessions = AttendanceSession.objects.filter(course=course, is_completed=True)
    enrollments = Enrollment.objects.filter(course=course, is_active=True).select_related('student__user')
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = [Paragraph(f'Attendance Report: {course.code} — {course.name}', styles['Title']),
                Spacer(1,12)]
    data = [['Student ID','Name','Total Classes','Attended','Percentage']]
    for enr in enrollments:
        s = enr.student
        total = sessions.count()
        att = AttendanceRecord.objects.filter(student=s,course=course,status='PRESENT').count()
        pct = f"{round(att/total*100,1)}%" if total else "0%"
        data.append([s.student_id, s.user.get_full_name(), str(total), str(att), pct])
    t = Table(data, colWidths=[80,160,80,70,70])
    t.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#0A4D68')),
        ('TEXTCOLOR',(0,0),(-1,0),colors.white),
        ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
        ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,colors.HexColor('#EEF2F7')]),
        ('GRID',(0,0),(-1,-1),0.5,colors.HexColor('#D0DAE8')),
        ('FONTSIZE',(0,0),(-1,-1),9),
        ('PADDING',(0,0),(-1,-1),6),
    ]))
    elements.append(t)
    doc.build(elements)
    buf.seek(0)
    resp = HttpResponse(buf, content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename="attendance_{course.code}.pdf"'
    return resp

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_excel(request, course_id):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return HttpResponse('openpyxl not installed', status=500)
    try:
        course = Course.objects.get(pk=course_id)
    except Course.DoesNotExist:
        return HttpResponse('Course not found', status=404)
    sessions = AttendanceSession.objects.filter(course=course, is_completed=True)
    enrollments = Enrollment.objects.filter(course=course, is_active=True).select_related('student__user')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = course.code
    header_fill = PatternFill(start_color='0A4D68', end_color='0A4D68', fill_type='solid')
    headers = ['Student ID','Name','Total Classes','Attended','Percentage']
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    for ri, enr in enumerate(enrollments, 2):
        s = enr.student
        total = sessions.count()
        att = AttendanceRecord.objects.filter(student=s,course=course,status='PRESENT').count()
        pct = round(att/total*100,1) if total else 0
        ws.append([s.student_id, s.user.get_full_name(), total, att, f"{pct}%"])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(buf, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="attendance_{course.code}.xlsx"'
    return resp
